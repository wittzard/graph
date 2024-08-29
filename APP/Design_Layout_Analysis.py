import streamlit as st
import pandas as pd
import networkx as nx
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
import math

def main():
    st.title("Design Layout Analysis")
    DesignFile = st.file_uploader("Upload Excel File", type=["xlsx"])
    DesignWorksheet = st.text_input("Enter Worksheet Name", value='Sheet1')
    DesignRange = st.text_input("Enter Range (e.g., A1:D10)", value='test')
    
    if DesignFile and DesignWorksheet and DesignRange:
        Layout = LoadFromExcel(DesignFile, DesignWorksheet, DesignRange)
        st.subheader("Layout")
        st.write(Layout)

        rows, columns = Layout.shape
        st.text(f"Rows: {rows}, Columns: {columns}")

        Labels = ApplyLabels(rows, columns)
        st.subheader("Labels")
        st.write(Labels)

        Adjacency = ConstructAdjacency(Layout, Labels, rows, columns)
        st.subheader("Adjacency Matrix")
        st.write(Adjacency)

        DesignGraph = CreateGraph(Adjacency)
        st.subheader("Design Graph")
        st.write(DesignGraph)

        Spots = SpotMatrix(Layout, Labels, rows, columns)
        st.subheader("Spot Matrix")
        st.write(Spots)

        Dists = Distances(Spots, DesignGraph)
        st.subheader("Distances")
        st.write(Dists)

        st.subheader("Design Graph Plot")
        PlotGraph(DesignGraph, Layout, Labels, rows, columns, Spots)

# Define all the necessary functions here
def LoadFromExcel(ExcelFile, Worksheet, Range):
    wb = load_workbook(filename=ExcelFile, read_only=True)
    ws = wb[Worksheet]
    dests = wb.defined_names[Range].destinations
    for title, coord in dests:
        min_col, min_row, max_col, max_row = range_boundaries(coord)
        data = ws.iter_rows(min_row, max_row, min_col, max_col, values_only=True)
    ExcelData = pd.DataFrame(data)
    return ExcelData

def ApplyLabels(rows, columns):
    currLabel = 1
    data = []
    for r in range(0, rows):
        column = []
        for c in range(0, columns):
            column.append(str(currLabel))
            currLabel += 1
        data.append(column)
    labels = pd.DataFrame(data)
    return labels

def ConstructAdjacency(Layout, Labels, rows, columns):
    node1 = []
    node2 = []
    weight = []
    cardinal = 5
    diagonal = 5 * (2**0.5)

    for row1 in range(rows):
        for column1 in range(columns):
            if Layout.loc[row1, column1].lower() in ('start', 'end', 'w'):
                for row2 in range(rows):
                    for column2 in range(columns):
                        if Layout.loc[row2, column2].lower() in ('start', 'end', 'w'):
                            if (abs(row2 - row1) <= 1) and (abs(column2 - column1) <= 1) and not (row1 == row2 and column1 == column2):
                                node1.append(Labels.loc[row1, column1])
                                node2.append(Labels.loc[row2, column2])
                                if (row1 == row2) or (column1 == column2):
                                    weight.append(cardinal)
                                else:
                                    weight.append(diagonal)

    edgelist = pd.DataFrame({
        'node1': node1,
        'node2': node2,
        'weights': weight
    })
    return edgelist

def CreateGraph(Adjacency):
    g = nx.DiGraph()
    for i, elrow in Adjacency.iterrows():
        g.add_edge(elrow['node1'], elrow['node2'], stepsize=round(elrow['weights'], 3))
    return g

def SpotMatrix(Layout, Labels, rows, columns):
    Cell = []
    Spot = []
    for row1 in range(0, rows):
        for column1 in range(0, columns):
            if Layout.loc[row1, column1].lower() == 'start':
                Cell.append(Labels.loc[row1, column1])
                Spot.append(Labels.loc[row1, column1])
    for row1 in range(0, rows):
        for column1 in range(0, columns):
            if Layout.loc[row1, column1].lower() == 'end':
                Cell.append(Labels.loc[row1, column1])
                Spot.append(Labels.loc[row1, column1])
    for row1 in range(rows):
        for column1 in range(columns):
            if Layout.loc[row1, column1].lower() == 's':
                cardinal_neighbors = [(row1 - 1, column1), (row1 + 1, column1), (row1, column1 - 1), (row1, column1 + 1)]
                for row2, column2 in cardinal_neighbors:
                    if Layout.loc[row2, column2].lower() == 'w':
                        Cell.append(Labels.loc[row1, column1])
                        Spot.append(Labels.loc[row2, column2])
    edgelist = pd.DataFrame({'Cell': Cell, 'Spot': Spot})
    return edgelist

def NetworkxPath(DesignGraph, input_node, output_node):
    PathLength = nx.shortest_path_length(DesignGraph, source=input_node, target=output_node, weight='stepsize')
    return PathLength

def Distances(Spots, DesignGraph):
    ShortestPaths = pd.DataFrame(index=range(len(Spots)), columns=range(len(Spots)))
    for row in range(len(Spots)):
        for column in range(len(Spots)):
            if row == column:
                ShortestPaths.at[row, column] = 0
            else:
                PathLength = NetworkxPath(DesignGraph, Spots.loc[row]['Spot'], Spots.loc[column]['Spot'])
                Decimals = round(math.log(1000, 10))
                ShortestPaths.at[row, column] = round(PathLength, Decimals)
                ShortestPaths.at[column, row] = round(PathLength, Decimals)
    return ShortestPaths

def PlotGraph(g, Layout, Labels, rows, columns, Spots):
    currName = 0
    for r in range(1, rows + 1):
        for c in range(1, columns + 1):
            currName += 1
            g.add_node(str(currName), pos=(c, rows - r))
    
    pos = nx.get_node_attributes(g, 'pos')
    NodeSize = 1500
    options1 = {'edgecolors': 'black', 'node_size': NodeSize, 'alpha': 1.0}
    options2 = {'node_size': 0}
    plt.figure(figsize=(rows * 1, columns * 1))
    Nodes1 = []
    Nodes2 = []
    NodeColourList1 = []
    ColorDict = {'w': '#D9D9D9', 'n': '#E7E4DB', 's': '#c00000', 'b': '#000000', 'o': '#F7C7AC',
                 'g': '#B5E6A2', 'p': '#F2CEEF', 'l': '#A6C9EC', 'm': '#D86DCD', 'spot': 'orange'}
    for row in range(0, rows):
        for column in range(0, columns):
            nodeColor = ColorDict[Layout.loc[row, column]]
            for CurrSpot in range(2, len(Spots)):
                if Labels.loc[row, column] == Spots.iloc[CurrSpot][1]:
                    nodeColor = ColorDict['spot']
            Nodes1.append(Labels.loc[row, column])
            NodeColourList1.append(nodeColor)
            if Layout.loc[row, column].lower() == 'b':
                Nodes2.append(Labels.loc[row, column])
    nx.draw(g, pos, nodelist=Nodes1, with_labels=True, font_color='Black', font_size=18, node_color=NodeColourList1, **options1)
    nx.draw(g.subgraph(Nodes2), pos, with_labels=True, font_color='White', font_size=18, **options2)
    nx.draw_networkx_edges(g, pos, node_size=NodeSize, arrowsize=20)
    SpotConnect = nx.DiGraph()
    for i, elrow in Spots.iterrows():
        SpotConnect.add_edge(elrow[0], elrow[1])
    nx.draw_networkx_edges(SpotConnect, pos, node_size=NodeSize, arrowsize=20, width=3.0, edge_color=ColorDict['spot'])
    st.pyplot(plt)

main()
