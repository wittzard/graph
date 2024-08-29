# Import dependencies
import networkx as nx
import pandas as pd
import numpy as np
import time as tm
import math as math
import matplotlib.pyplot as plt
import simpy
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from ortools.constraint_solver import routing_enums_pb2
from ortools.constraint_solver import pywrapcp
from matplotlib.animation import FuncAnimation
import streamlit as st
import io

# Title

if 'analysis_done' not in st.session_state:
    st.session_state.analysis_done = False

st.title("Design Layout Simulation")
st.header("Convert excel file to find graph")
# File uploader for the user to upload a file
uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

# Default file path
default_file = 'modified_map5.xlsx'

# Determine which file to use
if uploaded_file is not None:
    DesignFile = uploaded_file
else:
    DesignFile = default_file

# User inputs
DesignWorksheet = st.text_input("Enter Worksheet Name", value='Sheet1')
DesignRange = st.text_input("Enter Range (e.g., A1:D10)", value='test')
Avgwalk = st.number_input("Average walking speed (range 80-100 meter per min)", value=100, min_value=80, max_value=100, step=1)

#DesignFile = 'modified_map5.xlsx'  
#DesignWorksheet = 'Sheet1' 
#DesignRange = 'test'   
IntMultiplier = 1000  
#Avgwalk = 100 # meter per min

# Load layout from Excel file to pandas dataframe
def LoadFromExcel(ExcelFile, Worksheet, Range):
    wb = load_workbook(filename=ExcelFile, read_only=True)
    ws = wb[Worksheet]
    dests = wb.defined_names[Range].destinations
    for title, coord in dests:
        min_col, min_row, max_col, max_row = range_boundaries(coord)
        data = ws.iter_rows(min_row, max_row, min_col, max_col, values_only=True)
    ExcelData = pd.DataFrame(data)
    return ExcelData

# Create array of cell number labels.
# Cells are labelled from west-to-east and from north-to-south. row1.e. top-left = 1, bottom-right = n
def ApplyLabels(rows, columns):
    currLabel = 1
    data = []
    for r in range(0, rows):
        column = []
        for c in range(0, columns):
            column.append(str(currLabel))
            currLabel += 1
        data.append(column)
    labels = pd.DataFrame(data)
    return labels

# Build edgelist
def ConstructAdjacency(Layout, Labels, rows, columns):
    node1 = []
    node2 = []
    weight = []
    cardinal = 5
    diagoanal = 5*(2**0.5)

    for row1 in range(rows):
        for column1 in range(columns):
            if Layout.loc[row1,column1].lower() in ('start','end','w'):
                for row2 in range(rows):
                    for column2 in range(columns):
                        if Layout.loc[row2,column2].lower() in ('start','end','w'):
                            # check if adjacent
                            if (abs(row2-row1) <= 1) and (abs(column2-column1) <= 1) and not(row1 == row2 and column1 == column2):
                                node1.append(Labels.loc[row1,column1])
                                node2.append(Labels.loc[row2,column2])
                                # add weight
                                if (row1 == row2) or (column1 == column2):
                                    weight.append(cardinal)
                                else:
                                    weight.append(diagoanal)
    
    edgelist = pd.DataFrame({
        'node1': node1,
        'node2': node2,
        'weights':weight
    })
    return edgelist

# Create adjacency graph
def CreateGraph(Adjacency):
    g = nx.DiGraph()
    for i, elrow in Adjacency.iterrows():
        g.add_edge(elrow['node1'], elrow['node2'], stepsize=round(elrow['weights'],3))
    return g

def SpotMatrix(Layout, Labels, rows, columns):
    name = []
    Cell = []
    Spot = []
    row_column = []
    cartesian = []
    name_count = 0
    for row1 in range(rows):
        for column1 in range(columns):
            if Layout.loc[row1, column1].lower() == 's':
                cardinal_neighbors = [(row1-1, column1), # North
                                      (row1+1, column1), # South
                                      (row1, column1-1), # West
                                      (row1, column1+1)] # East
                for row2, column2 in cardinal_neighbors:
                    if 0 <= row2 < rows and 0 <= column2 < columns:  # Ensure within bounds
                        if Layout.loc[row2, column2].lower() == 'w':
                            Cell.append(Labels.loc[row1, column1])
                            Spot.append(Labels.loc[row2, column2])
                            row_column.append((row2, column2))  # Add row_column coordinates
                            cartesian.append((column2,-row2))
                            name.append(name_count)
                            name_count += 1
                            
    edgelist = pd.DataFrame({'name': name, 'Cell': Cell, 'Spot': Spot, 'row_column': row_column, 'Cartesian': cartesian})
    #just swap depot @ 7 to 0
    temp_row = edgelist.loc[0, ['Cell', 'Spot', 'row_column', 'Cartesian']].copy()
    edgelist.loc[0, ['Cell', 'Spot', 'row_column', 'Cartesian']] = edgelist.loc[7, ['Cell', 'Spot', 'row_column', 'Cartesian']]
    edgelist.loc[7, ['Cell', 'Spot', 'row_column', 'Cartesian']] = temp_row
    return edgelist

# Shortest distance between two spots, using networkx
def NetworkxPath(DesignGraph, input_node, output_node):
    PathLength = nx.shortest_path_length(DesignGraph, source = input_node, target = output_node, weight='stepsize')
    return PathLength

# Create matrix of shortest paths between all spot cells
def Distances(Spots, DesignGraph):
    ShortestPaths = pd.DataFrame(index=range(len(Spots)), columns=range(len(Spots)))
    for row in range(len(Spots)):
        for column in range(len(Spots)):
            if row == column:
                ShortestPaths.at[row,column] = 0
            else:
                PathLength = NetworkxPath(DesignGraph, Spots.loc[row]['Spot'], Spots.loc[column]['Spot'])
                Decimals = round(math.log(IntMultiplier,10))
                ShortestPaths.at[row, column] = round(PathLength, Decimals)   # Upper-right triangle of matrix
                ShortestPaths.at[column, row] = round(PathLength, Decimals)   # Lower-left triangle of matrix
    dist = ShortestPaths.astype(int)
    dist = dist / Avgwalk
    return dist

def PlotGraph(g, Layout, Labels, rows, columns, Spots):
    currName = 0
    for r in range(1, rows + 1):
        for c in range(1, columns + 1):
            currName += 1
            g.add_node(str(currName), pos=(c, rows - r))
    
    pos = nx.get_node_attributes(g, 'pos')
    NodeSize = 1500
    options1 = {'edgecolors': 'black', 'node_size': NodeSize, 'alpha': 1.0}
    options2 = {'node_size': 0}
    plt.figure(figsize=(rows * 1, columns * 1))
    Nodes1 = []
    Nodes2 = []
    NodeColourList1 = []
    ColorDict = {'w': '#D9D9D9', 'n': '#E7E4DB', 's': '#c00000', 'b': '#000000', 'o': '#F7C7AC',
                 'g': '#B5E6A2', 'p': '#F2CEEF', 'l': '#A6C9EC', 'm': '#D86DCD', 'spot': 'orange'}
    for row in range(0, rows):
        for column in range(0, columns):
            nodeColor = ColorDict[Layout.loc[row, column]]
            for CurrSpot in range(2, len(Spots)):
                if Labels.loc[row, column] == Spots.iloc[CurrSpot][1]:
                    nodeColor = ColorDict['spot']
            Nodes1.append(Labels.loc[row, column])
            NodeColourList1.append(nodeColor)
            if Layout.loc[row, column].lower() == 'b':
                Nodes2.append(Labels.loc[row, column])
    nx.draw(g, pos, nodelist=Nodes1, with_labels=True, font_color='Black', font_size=18, node_color=NodeColourList1, **options1)
    nx.draw(g.subgraph(Nodes2), pos, with_labels=True, font_color='White', font_size=18, **options2)
    nx.draw_networkx_edges(g, pos, node_size=NodeSize, arrowsize=20)
    SpotConnect = nx.DiGraph()
    for i, elrow in Spots.iterrows():
        SpotConnect.add_edge(elrow[0], elrow[1])
    nx.draw_networkx_edges(SpotConnect, pos, node_size=NodeSize, arrowsize=20, width=3.0, edge_color=ColorDict['spot'])
    st.pyplot(plt)

def plot_spots_with_cartesian(edgelist, distance_matrix):
    G = nx.Graph()

    # Add nodes to the graph
    G.add_nodes_from(edgelist['name'])

    # Extract depot and elements
    depot = edgelist.iloc[0]
    elements = edgelist.drop(0)

    # Create a dictionary for node positions
    pos = {row['name']: row['Cartesian'] for _, row in edgelist.iterrows()}

    # Add fully connected edges using the distance matrix
    for i in range(len(distance_matrix)):
        for j in range(i + 1, len(distance_matrix)):
            node_i = edgelist.iloc[i]['name']
            node_j = edgelist.iloc[j]['name']
            distance = distance_matrix.iloc[i, j]
            G.add_edge(node_i, node_j, weight=distance)

    # Plot settings
    plt.figure(figsize=(10, 10))  # Increase figure size

    # Plot the depot node in red
    nx.draw_networkx_nodes(G, pos, nodelist=[depot['name']], node_color='red', node_size=500)

    # Plot the other spot nodes in blue
    nx.draw_networkx_nodes(G, pos, nodelist=elements['name'], node_color='blue', node_size=500)

    # Add labels to the nodes
    nx.draw_networkx_labels(G, pos, font_size=12, font_weight='bold')

    # Draw the edges with a thicker line width
    nx.draw_networkx_edges(G, pos, edge_color='gray', width=2)

    # Optionally, add edge labels for distances
    edge_labels = nx.get_edge_attributes(G, 'weight')
    nx.draw_networkx_edge_labels(G, pos, edge_labels=edge_labels, font_size=8)

    buf = io.BytesIO()
    plt.savefig(buf, format='png')
    buf.seek(0)

    # Display the plot in Streamlit
    st.image(buf, caption='Node graph', use_column_width=True)

    #return G, pos

def run():
    Layout = LoadFromExcel(DesignFile,DesignWorksheet,DesignRange)   
    print(Layout)

    rows,columns = Layout.shape
    print(f"\nrows {rows} columns {columns}")

    Labels = ApplyLabels(rows, columns)
    print(Labels)
    
    Adjacency = ConstructAdjacency(Layout, Labels, rows, columns)
    print(Adjacency)

    DesignGraph = CreateGraph(Adjacency)
    print(DesignGraph)

    Spots = SpotMatrix(Layout, Labels, rows, columns)
    print(Spots)

    st.write("Node graph from excel")
    PlotGraph(DesignGraph, Layout, Labels, rows, columns, Spots)

    timematrix = Distances(Spots, DesignGraph) # multiply with walk factor
    print(timematrix)

    st.write("Node graph from time matrix")
    plot_spots_with_cartesian(Spots, timematrix)
    
    return Spots, timematrix
# Title
st.header("Simulation Selector")

# Choose simulation type
simulation_type = st.selectbox("Choose a simulation:", ["Vehicle Routing Problem (VRP)", "Traveling Salesman Problem (TSP)"])

# Display content based on selected simulation type
if simulation_type == "Vehicle Routing Problem (VRP)":
    #st.write("You have selected to simulate the Vehicle Routing Problem (VRP).")
    num_vehicles = st.number_input("Number of Vehicles", value=4, min_value=2, max_value=4, step=1, help="range 2-4")
    max_time_per_vehicle = st.number_input("Max time per vehicle (minutes)", value=30, min_value=30, max_value=100, step=1)

elif simulation_type == "Traveling Salesman Problem (TSP)":
    #st.write("You have selected to simulate the Traveling Salesman Problem (TSP).")
    # For TSP, num_vehicles is always 1, so no need for a number input
    st.write("Number of Vehicles is set to 1 for TSP.")
    num_vehicles = 1
    max_time_per_vehicle = st.number_input("Max time per vehicle (minutes)", value=30, min_value=30, max_value=100, step=1)

st.selectbox("Choose a Strategy:", ["NearestNeighbor"])

if st.button('Start Simulation',key='start_simulation'):
    st.write("Upload Excel file successfuly!")
    st.write("Running code...")
    Spots, timematrix = run()  
    pots_df = Spots  # DataFrame with spot locations, for animation
    time_matrix = timematrix  # Time matrix between nodes
    depot_index = 0
    
    class Customer:
        def __init__(self, id):
            self.id = id

    class Vehicle:
        def __init__(self, env, id, depot_index, max_time):
            self.env = env
            self.id = id
            self.current_location = depot_index
            self.route = []
            self.time_spent = 0
            self.max_time = max_time
            self.finished = False
            self.served_customers = []

        def travel_to(self, customer_index):
            travel_time = time_matrix[self.current_location][customer_index]
            yield self.env.timeout(travel_time)
            self.time_spent += travel_time
            self.current_location = customer_index
            print(f"Vehicle {self.id} traveled to customer {customer_index} at time {self.env.now}")

        def serve_customer(self, customer):
            service_time = 1  # Fixed service time
            yield self.env.timeout(service_time)
            self.time_spent += service_time
            print(f"Vehicle {self.id} served customer {customer} at time {self.env.now}")

        def run(self):
            while not self.finished:
                if not self.route:
                    yield self.env.timeout(1)  # Wait for new assignments
                    continue

                next_customer_index = self.route.pop(0)
                # Travel to the next customer
                yield self.env.process(self.travel_to(next_customer_index))
                # Serve the next customer
                yield self.env.process(self.serve_customer(next_customer_index))
                # Record served customer
                self.served_customers.append(next_customer_index)
                # Return to depot if necessary
                if not self.route or self.time_spent + time_matrix[self.current_location][depot_index] > self.max_time:
                    yield self.env.process(self.travel_to(depot_index))
                    print(f"Vehicle {self.id} returned to depot at time {self.env.now}")
                    self.finished = True

    class NearestNeighborStrategy:
        def __init__(self, time_matrix, depot_index):
            self.time_matrix = time_matrix
            self.depot_index = depot_index

        def choose_route(self, vehicle, unvisited_customers):
            current_location = vehicle.current_location
            route = []
            while unvisited_customers:
                nearest_customer = min(
                    unvisited_customers,
                    key=lambda cust: self.time_matrix[current_location][cust]
                )
                route.append(nearest_customer)
                unvisited_customers.remove(nearest_customer)
                current_location = nearest_customer
            return route

    def vrptw_simulation(env):
        vehicles = [Vehicle(env, i, depot_index, max_time_per_vehicle) for i in range(num_vehicles)]
        strategy = NearestNeighborStrategy(time_matrix, depot_index)
        
        # Create a list of unvisited customers
        unvisited_customers = [c for c in range(1, len(time_matrix))]
        
        # Assign routes to each vehicle using the Nearest Neighbor strategy
        for vehicle in vehicles:
            vehicle.route = strategy.choose_route(vehicle, unvisited_customers.copy())
            env.process(vehicle.run())

        while any(not v.finished for v in vehicles):
            yield env.timeout(1)

        all_results = {
            'served_customers': {v.id: v.served_customers for v in vehicles}
        }
        return all_results
   
    # Create simulation environment
    env = simpy.Environment()
    results_process = env.process(vrptw_simulation(env))
    env.run()

    # Collect and display reslts
    results = results_process.value
    for vehicle_id, served_customers in results['served_customers'].items():
                print(f"Vehicle {vehicle_id} served customers: {served_customers}")

    routes = results['served_customers']
    for vehicle_id, route in routes.items():
                route.insert(0, 0)  # Add 0 at the start
                route.append(0)     # Add 0 at the end
            # Print the updated routes
        

    G = nx.Graph()
    # Add nodes to the graph
    G.add_nodes_from(Spots['name'])
    # Create a dictionary for node positions
    pos = {row['name']: row['Cartesian'] for _, row in Spots.iterrows()}

    # Add fully connected edges using the timematrixtance matrix
    for i in range(len(timematrix)):
        for j in range(i + 1, len(timematrix)):
            node_i = Spots.iloc[i]['name']
            node_j = Spots.iloc[j]['name']
            timematrixtance = timematrix.iloc[i, j]
            G.add_edge(node_i, node_j, weight=timematrixtance)

    fig, ax = plt.subplots(figsize=(10, 10))

    nx.draw_networkx_nodes(G, pos, ax=ax, node_color='blue', node_size=500)
    # Highlight the depot node
    nx.draw_networkx_nodes(G, pos, ax=ax, nodelist=[Spots.iloc[0]['name']], node_color='red', node_size=700)
    # Draw all edges
    nx.draw_networkx_edges(G, pos, ax=ax, edge_color='gray', width=1)

    colors = ['green', 'orange', 'red', 'purple']  # Define colors for each vehicle

    def update(frame):
        ax.clear()
        nx.draw_networkx_nodes(G, pos, ax=ax, node_color='blue', node_size=500)
        nx.draw_networkx_nodes(G, pos, ax=ax, nodelist=[Spots.iloc[0]['name']], node_color='red', node_size=700)
        nx.draw_networkx_edges(G, pos, ax=ax, edge_color='gray', width=1)

        # Determine the current vehicle to display based on the frame
        num_vehicles = len(routes)
        vehicle_id = (frame // (len(max(routes.values(), key=len))) % num_vehicles)

        # Get the route for the current vehicle
        route = routes[vehicle_id]
        path_length = len(route)
        frame_index = min(frame % path_length, path_length - 1)
        path = route[:frame_index + 1]

        # Draw the route for the current vehicle
        if len(path) > 1:
            nx.draw_networkx_edges(G, pos, edgelist=[(path[i], path[i + 1]) for i in range(len(path) - 1)],
                                ax=ax, edge_color=colors[vehicle_id % len(colors)], width=2, alpha=0.7)

        # Plot the vehicle's position
        if path:
            nx.draw_networkx_nodes(G, pos, nodelist=[path[-1]], node_color=colors[vehicle_id % len(colors)],
                                node_size=500, label=f'Vehicle {vehicle_id}')

        # Add a legend for vehicle colors
        handles = [plt.Line2D([0], [0], marker='o', color='w', markerfacecolor=color, markersize=10, label=f'Vehicle {i}') for i, color in enumerate(colors)]
        ax.legend(handles=handles)

    # Calculate the number of frames needed for the animation
    max_frames = len(max(routes.values(), key=len)) * len(routes)

    fig, ax = plt.subplots(figsize=(10, 10))
    ani = FuncAnimation(fig, update, frames=max_frames, repeat=False, interval=1000)

    # Save or show the animation
    ani.save('Layout_simulation_sequence.gif', writer='pillow')

    st.image('Layout_simulation_sequence.gif')




